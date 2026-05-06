import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuração dos meses para análise.
# Modifique esta lista com os meses que deseja incluir na análise.
MESES_ANALISADOS = ['Jan/2026', 'Fev/2026', 'Mar/2026', 'Abr/2026', 'Mai/2026']
ARQUIVO_ENTRADA = 'TO Hitlist.xlsx'
ARQUIVO_SAIDA = 'TO Hitlist.xlsx'

def map_month_to_anomes(month_str):
    parts = month_str.split(' ')
    if len(parts) == 2:
        month_num = parts[1].split('/')[0]
        year_short = parts[1].split('/')[1]

        month_map = {
            '01': 'Jan', '02': 'Fev', '03': 'Mar', '04': 'Abr',
            '05': 'Mai', '06': 'Jun', '07': 'Jul', '08': 'Ago',
            '09': 'Set', '10': 'Out', '11': 'Nov', '12': 'Dez'
        }

        month_name = month_map.get(month_num)
        if month_name:
            return f"{month_name}/20{year_short}"
    return month_str

print(f"Lendo dados do arquivo '{ARQUIVO_ENTRADA}'...")

# Ler as planilhas
df_csat = pd.read_excel(ARQUIVO_ENTRADA, sheet_name='csat-sqs')
df_abs = pd.read_excel(ARQUIVO_ENTRADA, sheet_name='abs')

# Padronizar datas e nomes
df_csat['Month'] = df_csat['Month'].apply(map_month_to_anomes)
df_csat['AGENT'] = df_csat['AGENT'].astype(str).str.strip().str.upper()
df_abs['Nome'] = df_abs['Nome'].astype(str).str.strip().str.upper()

# Filtrar pelos meses selecionados
df_csat_filtered = df_csat[df_csat['Month'].str.title().isin([m.title() for m in MESES_ANALISADOS])].copy()
df_abs_filtered = df_abs[df_abs['AnoMes'].str.title().isin([m.title() for m in MESES_ANALISADOS])].copy()

if df_csat_filtered.empty:
    print(f"Aviso: Nenhum dado de CSAT/SQS encontrado para os meses: {MESES_ANALISADOS}")
    exit()

# Calcular Médias Globais do período
global_csat_answered = df_csat_filtered[df_csat_filtered.columns[3]].sum()
global_csat_sum = (df_csat_filtered['% CSAT'] * df_csat_filtered[df_csat_filtered.columns[3]]).sum()
global_csat = global_csat_sum / global_csat_answered if global_csat_answered > 0 else 0

global_sqs = df_csat_filtered['% SQS Agent Level'].mean()

# Agregar dados por agente (CSAT e SQS)
def weighted_avg(values, weights):
    s = weights.sum()
    if s == 0:
        return 0
    return (values * weights).sum() / s

agent_csat = df_csat_filtered.groupby('AGENT').apply(
    lambda x: pd.Series({
        df_csat_filtered.columns[3]: x[df_csat_filtered.columns[3]].sum(),
        '% CSAT': weighted_avg(x['% CSAT'], x[df_csat_filtered.columns[3]]),
        'QMS Submission': x['QMS Submission'].sum(),
        '% SQS Agent Level': x['% SQS Agent Level'].mean()
    })
).reset_index()

# Agregar dados de ABS por agente
agent_abs = df_abs_filtered.groupby('Nome').agg({
    'ABS Total': 'mean'
}).reset_index().rename(columns={'Nome': 'AGENT'})

# Juntar as bases
merged_data = pd.merge(agent_csat, agent_abs, on='AGENT', how='outer')
merged_data = merged_data.fillna(0)

# ==========================================
# 1. CÁLCULOS REAIS
# ==========================================
merged_data['CSAT Impact'] = (merged_data['% CSAT'] - global_csat) * merged_data[df_csat_filtered.columns[3]]
merged_data['SQS Impact'] = (merged_data['% SQS Agent Level'] - global_sqs) * merged_data['QMS Submission']
merged_data['Total Real Impact'] = merged_data['CSAT Impact'] + merged_data['SQS Impact']

# ==========================================
# 2. CÁLCULOS PROJETADOS (Meta ABS = 5%)
# ==========================================
# Proporção = (95% de presença desejada) / (Presença real)
merged_data['Proportion'] = 0.0
mask = merged_data['ABS Total'] < 1.0
merged_data.loc[mask, 'Proportion'] = 0.95 / (1.0 - merged_data.loc[mask, 'ABS Total'])

merged_data['Proj Answered CSAT'] = merged_data[df_csat_filtered.columns[3]] * merged_data['Proportion']
num_months = len(MESES_ANALISADOS)
max_qms_projected = 8 * num_months
merged_data['Proj QMS Submission'] = merged_data['QMS Submission'] * merged_data['Proportion']
merged_data['Proj QMS Submission'] = merged_data['Proj QMS Submission'].clip(upper=max_qms_projected)

merged_data['Proj CSAT Impact'] = (merged_data['% CSAT'] - global_csat) * merged_data['Proj Answered CSAT']
merged_data['Proj SQS Impact'] = (merged_data['% SQS Agent Level'] - global_sqs) * merged_data['Proj QMS Submission']
merged_data['Total Proj Impact'] = merged_data['Proj CSAT Impact'] + merged_data['Proj SQS Impact']

# ==========================================
# 3. GERAR OS RANKINGS
# ==========================================
# Função auxiliar para ordenar
def sort_ranking(df, sort_col, cols_to_keep, ascending=True):
    return df.sort_values(by=sort_col, ascending=ascending)[cols_to_keep].copy()

# A. Impacto Apenas CSAT
rank_csat = sort_ranking(merged_data, 'CSAT Impact', ['AGENT', '% CSAT', df_csat_filtered.columns[3], 'CSAT Impact'])
# B. Impacto Apenas SQS
rank_sqs = sort_ranking(merged_data, 'SQS Impact', ['AGENT', '% SQS Agent Level', 'QMS Submission', 'SQS Impact'])
# C. Apenas ABS (Do maior para o menor = piores primeiro)
rank_abs = sort_ranking(merged_data, 'ABS Total', ['AGENT', 'ABS Total'], ascending=False)
# D. Classificação 1: Impacto Total Real (CSAT + SQS)
rank_real_total = sort_ranking(merged_data, 'Total Real Impact', ['AGENT', 'CSAT Impact', 'SQS Impact', 'Total Real Impact'])
# E. Classificação 2: Impacto Total Real c/ ABS Visualização
rank_real_abs = sort_ranking(merged_data, 'Total Real Impact', ['AGENT', 'Total Real Impact', 'ABS Total'])
# F. Classificação 3: Impacto Total Projetado
rank_proj_total = sort_ranking(merged_data, 'Total Proj Impact', ['AGENT', 'Proj CSAT Impact', 'Proj SQS Impact', 'Total Proj Impact'])
# G. Classificação 4: Impacto Total Projetado c/ ABS Visualização
rank_proj_abs = sort_ranking(merged_data, 'Total Proj Impact', ['AGENT', 'Total Proj Impact', 'ABS Total'])

# ==========================================
# 4. SALVAR NO EXCEL
# ==========================================
print(f"Escrevendo os resultados na aba 'Top Ofenders'...")

# Vamos usar openpyxl para não sobrescrever as outras abas sem querer
try:
    book = openpyxl.load_workbook(ARQUIVO_SAIDA)
except FileNotFoundError:
    print(f"Erro: Arquivo '{ARQUIVO_SAIDA}' não encontrado.")
    exit()

if 'Top Ofenders' in book.sheetnames:
    del book['Top Ofenders']

sheet = book.create_sheet('Top Ofenders')

def append_df_to_sheet(sheet, df, title, start_row, start_col):
    sheet.cell(row=start_row, column=start_col, value=title).font = openpyxl.styles.Font(bold=True)
    start_row += 1
    # Headers
    for c_idx, col_name in enumerate(df.columns):
        sheet.cell(row=start_row, column=start_col + c_idx, value=col_name).font = openpyxl.styles.Font(bold=True)

    # Dados
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False)):
        for c_idx, value in enumerate(row):
            sheet.cell(row=start_row + 1 + r_idx, column=start_col + c_idx, value=value)

# Adicionar textos globais
sheet.cell(row=1, column=1, value="MÉDIAS GLOBAIS DO PERÍODO SELECIONADO:").font = openpyxl.styles.Font(bold=True)
sheet.cell(row=2, column=1, value=f"% CSAT Global: {global_csat:.2%}")
sheet.cell(row=3, column=1, value=f"% SQS Global: {global_sqs:.2%}")
sheet.cell(row=4, column=1, value=f"Meses Analisados: {', '.join(MESES_ANALISADOS)}")

# Posicionar as tabelas
append_df_to_sheet(sheet, rank_csat, "1. Ranking Individual: CSAT Impact", 6, 1)
append_df_to_sheet(sheet, rank_sqs, "2. Ranking Individual: SQS Impact", 6, 6)
append_df_to_sheet(sheet, rank_abs, "3. Ranking Individual: ABS", 6, 11)

append_df_to_sheet(sheet, rank_real_total, "4. Classificação Real (CSAT + SQS)", 20, 1)
append_df_to_sheet(sheet, rank_real_abs, "5. Classificação Real c/ Visualização ABS", 20, 6)
append_df_to_sheet(sheet, rank_proj_total, "6. Classificação Projetada (Alvo 5% ABS)", 20, 10)
append_df_to_sheet(sheet, rank_proj_abs, "7. Classificação Proj c/ Visualização ABS", 20, 15)

# Formatar os dados
for row in sheet.iter_rows(min_row=6):
    for cell in row:
        if isinstance(cell.value, float):
            # Formatar ABS e CSAT como percentual onde fizer sentido
            # Para simplificar o visual, arredonda float
            cell.number_format = '0.00'

book.save(ARQUIVO_SAIDA)
print("Concluído! Abra o arquivo e confira a aba 'Top Ofenders'.")
